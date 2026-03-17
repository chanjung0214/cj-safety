import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import os
import io

class SafetyEduReportGenerator:
    # ⭐ [수정1] 이제 양식 파일이 1개로 합쳐졌으니, 템플릿도 1개만 받습니다!
    def __init__(self, integrated_template):
        self.integrated_template = integrated_template

    # ⭐ [수정2] 교육일지와 사진대지 작업을 하나로 합친 통합 프로세스입니다!
    def process_integrated_report(self, sheet_numbers, site_name, date_str, instructor, count, time_range, content_str, photo_paths):
        # 1. 42개+사진대지가 포함된 통합 템플릿 열기
        wb = openpyxl.load_workbook(self.integrated_template)
        
        # 날짜 포맷팅
        date_korean = f"20{date_str[:2]}년 {date_str[2:4]}월 {date_str[4:]}일"
        date_dash = f"20{date_str[:2]}-{date_str[2:4]}-{date_str[4:]}"
        
        # 2. 교육일지 시트들에 내용 채우기
        for num in sheet_numbers:
            if num in wb.sheetnames:
                ws = wb[num]
                ws['C4'] = site_name    # 현장명
                ws['A3'] = date_korean  # 교육일자
                ws['I4'] = instructor   # 실시자
                ws['C6'] = f"대상( {count} 名)      참석( {count} 名)      미실시( 0 名)"
                ws['I7'] = time_range   # 교육시간
        
        # 3. '사진대지' 시트에 사진 및 내용 채우기
        if '사진대지' in wb.sheetnames:
            ws_photo = wb['사진대지']
            # 시트 이름 변경은 생략하거나 그대로 둡니다 (삭제 방지를 위해 '사진대지' 유지)
            
            positions = [
                {'img': 'A2',  'date': 'B21', 'content': 'E21'},
                {'img': 'A23', 'date': 'B42', 'content': 'E42'},
                {'img': 'I2',  'date': 'J21', 'content': 'M21'},
                {'img': 'I23', 'date': 'J42', 'content': 'M42'}
            ]
            
            for idx, img_path in enumerate(photo_paths):
                if idx >= 4:
                    break
                
                pos = positions[idx]
                ws_photo[pos['date']] = date_dash
                ws_photo[pos['content']] = content_str
                
                # 이미지 리사이징 및 삽입 (대리님이 수정한 610, 400 사이즈 적용!)
                if img_path and os.path.exists(img_path):
                    img = PILImage.open(img_path)
                    img = img.resize((610, 400), PILImage.Resampling.LANCZOS)
                    
                    temp_img_path = f"temp_{idx}.png"
                    img.save(temp_img_path)
                    
                    excel_img = ExcelImage(temp_img_path)
                    ws_photo.add_image(excel_img, pos['img'])
        
        # 4. 🧹 살려둘 시트만 남기고 쩌리 시트 삭제
        # 대리님이 기존에 쓰시던 '서명지', '서명부'도 안전하게 살려둡니다!
        sheets_to_keep = sheet_numbers + ["서명지", "서명부", "사진대지"] 
        for sheet_name in wb.sheetnames:
            if sheet_name not in sheets_to_keep:
                del wb[sheet_name]
                
      # 5. 엑셀 최종 저장 먼저! (이때 엑셀이 임시 사진을 진짜로 빨아들입니다)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 6. 저장이 무사히 끝난 후 임시 이미지 파일 청소
        for f in os.listdir('.'):
            if f.startswith('temp_') and f.endswith('.png'):
                try:
                    os.remove(f)
                except Exception:
                    pass
                
        return output

    # ⭐ 마스터 DB 업데이트 코드는 대리님이 짜두신 완벽한 상태 그대로 둡니다!
    def update_master_db(self, date_str, instructor, count, time_range, selected_names):
        db_path = "안전교육_관리대장.xlsx"
        date_korean = f"20{date_str[:2]}년 {date_str[2:4]}월 {date_str[4:]}일"
        
        if not os.path.exists(db_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "교육대장"
            ws.append(["교육일자", "실시자", "교육인원(명)", "교육시간", "교육항목"])
        else:
            wb = openpyxl.load_workbook(db_path)
            ws = wb.active
            
        items_str = ", ".join(selected_names)
        ws.append([date_korean, instructor, count, time_range, items_str])
        wb.save(db_path)